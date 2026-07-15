"""Extract Radical Red v4.1 boss teams from the public trainer workbooks.

The source sheets use repeated visual blocks rather than conventional tables.
Each block has its trainer label in column C and up to six Pokemon beginning in
columns E, J, O, T, Y and AD.  Keeping the extractor in the repository makes
the generated guide data reproducible when the documentation is updated.
"""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCES = {
    "default": ROOT / "work" / "radical-red-default-bosses-v4.1.xlsx",
    "hardcore": ROOT / "work" / "radical-red-hardcore-bosses-v4.1.xlsx",
}
OUTPUT_JSON = ROOT / "data" / "battle-data.json"
OUTPUT_JS = ROOT / "data" / "battle-data.js"
TEAM_SHEETS = [
    "Kanto Leaders",
    "Kanto Rematch",
    "Johto Leaders",
    "Rivals",
    "Team Rocket",
    "Mini Bosses",
    "Optional Bosses",
    "Indigo League",
    "Postgame",
]
POKEMON_COLUMNS = (5, 10, 15, 20, 25, 30)
STAT_NAMES = ("hp", "attack", "defense", "spAttack", "spDefense", "speed")


def text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def is_level(value) -> bool:
    if isinstance(value, (int, float)):
        return 1 <= value <= 100
    value = text(value).lower()
    return bool(value and ("lv" in value or "level" in value))


def battle_rows(sheet) -> list[int]:
    rows = []
    for row in range(1, sheet.max_row + 1):
        trainer = text(sheet.cell(row, 3).value)
        if not trainer:
            continue
        species = [
            text(sheet.cell(row, col).value)
            for col in POKEMON_COLUMNS
            if text(sheet.cell(row, col).value)
            and is_level(sheet.cell(row + 1, col).value)
        ]
        if species:
            rows.append(row)
    return rows


def extract_pokemon(sheet, row: int, col: int) -> dict | None:
    name = text(sheet.cell(row, col).value)
    level = sheet.cell(row + 1, col).value
    if not name or not is_level(level):
        return None

    base_stats = {}
    evs = {}
    for offset, key in enumerate(STAT_NAMES, start=13):
        base = sheet.cell(row + offset, col + 1).value
        ev = sheet.cell(row + offset, col + 3).value
        if isinstance(base, (int, float)):
            base_stats[key] = int(base)
        if isinstance(ev, (int, float)) and ev:
            evs[key] = int(ev)

    moves = [
        text(sheet.cell(row + offset, col).value)
        for offset in range(7, 11)
    ]
    moves = [move for move in moves if move and move != "-"]
    speed_stat = sheet.cell(row + 19, col + 3).value

    result = {
        "name": name,
        "level": int(level) if isinstance(level, (int, float)) else text(level),
        "nature": text(sheet.cell(row + 4, col).value),
        "ability": text(sheet.cell(row + 5, col).value),
        "item": text(sheet.cell(row + 6, col).value),
        "moves": moves,
    }
    if len(base_stats) == 6:
        result["baseStats"] = base_stats
        result["bst"] = sum(base_stats.values())
    if evs:
        result["evs"] = evs
    if isinstance(speed_stat, (int, float)):
        result["speedStat"] = int(speed_stat)
    return result


def extract_mode(path: Path, mode: str) -> list[dict]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    battles = []
    for sheet_name in TEAM_SHEETS:
        sheet = workbook[sheet_name]
        rows = battle_rows(sheet)
        labels = [text(sheet.cell(row, 3).value) for row in rows]
        totals = Counter(labels)
        occurrences = defaultdict(int)
        for row, label in zip(rows, labels):
            occurrences[label] += 1
            team = [extract_pokemon(sheet, row, col) for col in POKEMON_COLUMNS]
            team = [pokemon for pokemon in team if pokemon]
            notes = []
            for note_row in range(max(1, row - 4), row):
                for col in range(1, sheet.max_column + 1):
                    note = text(sheet.cell(note_row, col).value)
                    if note and any(token in note.upper() for token in ("BATTLE EFFECT", "BATTLE TYPE", "PERMANENT", "TERRAIN", "WEATHER")):
                        if note not in notes:
                            notes.append(note)
            numeric_levels = [member["level"] for member in team if isinstance(member["level"], int)]
            battle = {
                    "id": f"{mode}-{re.sub(r'[^a-z0-9]+', '-', sheet_name.lower()).strip('-')}-{row}",
                    "mode": mode,
                    "category": sheet_name,
                    "trainer": label,
                    "variant": occurrences[label] if totals[label] > 1 else None,
                    "documentOrder": len(battles) + 1,
                    "team": team,
                    "levelMin": min(numeric_levels) if numeric_levels else None,
                    "levelMax": max(numeric_levels) if numeric_levels else None,
                }
            if notes:
                battle["notes"] = notes
            battles.append(battle)
    return battles


def main() -> None:
    missing = [str(path) for path in SOURCES.values() if not path.exists()]
    if missing:
        raise SystemExit("Missing source workbook(s): " + ", ".join(missing))
    battles = []
    for mode, path in SOURCES.items():
        battles.extend(extract_mode(path, mode))
    payload = {
        "meta": {
            "version": "4.1",
            "sources": {
                "default": "https://docs.google.com/spreadsheets/d/1ES8L4OzeJ8rCuMWFNvrDaZKArqR7Vys2ytFxjx2pbwE/edit",
                "hardcore": "https://docs.google.com/spreadsheets/d/1jDbKFA30xo8csPHZNLtsmqs781bW_Xb9mKoPYyE6KK8/edit",
            },
        },
        "battles": battles,
    }
    encoded = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    OUTPUT_JSON.write_text(encoded, encoding="utf-8")
    OUTPUT_JS.write_text("window.BATTLE_DATA=" + encoded + ";\n", encoding="utf-8")
    counts = Counter((battle["mode"], battle["category"]) for battle in battles)
    print(f"Wrote {len(battles)} battle variants")
    for key, count in sorted(counts.items()):
        print(f"  {key[0]:8} {key[1]:18} {count:3}")


if __name__ == "__main__":
    main()
