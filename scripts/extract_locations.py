import json
from datetime import datetime
from pathlib import Path

import openpyxl


SOURCE = Path(r"C:\Users\james\Downloads\Pokémon Locations & Raid Dens v4.1 - Radical Red.xlsx")
OUTPUT = Path("work/grass-cave-locations.json")


def level_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return f"{value.month}-{value.day}"
    return str(value).replace("Lv. ", "").strip()


def encounter_block(sheet, start_col, header_row, first_row, last_row, period):
    name = sheet.cell(header_row, start_col + 1).value
    if not isinstance(name, str) or not name.strip():
        return None
    encounters = []
    for row in range(first_row, last_row + 1):
        rarity = sheet.cell(row, start_col).value
        pokemon = sheet.cell(row, start_col + 2).value
        level = sheet.cell(row, start_col + 3).value
        if not pokemon:
            continue
        encounters.append({
            "pokemon": str(pokemon).strip(),
            "rarity": round(float(rarity) * 100, 2) if isinstance(rarity, (int, float)) else None,
            "level": level_text(level),
        })
    return {"name": name.strip().title(), "period": period, "encounters": encounters}


def main():
    workbook = openpyxl.load_workbook(SOURCE, data_only=True, read_only=False)
    sheet = workbook["Grass & Caves"]
    locations = {}
    # Every location occupies five columns. The first table is day; the vertically
    # aligned table below it is night, matching the workbook's visual convention.
    for start_col in range(3, sheet.max_column + 1, 5):
        for block in (
            encounter_block(sheet, start_col, 3, 5, 16, "day"),
            encounter_block(sheet, start_col, 18, 20, 31, "night"),
        ):
            if not block:
                continue
            key = block["name"].casefold()
            item = locations.setdefault(key, {"name": block["name"], "day": [], "night": []})
            item[block["period"]] = block["encounters"]
    payload = sorted(locations.values(), key=lambda item: item["name"])
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Extracted {len(payload)} locations and {sum(len(x['day']) + len(x['night']) for x in payload)} encounter rows")


if __name__ == "__main__":
    main()
