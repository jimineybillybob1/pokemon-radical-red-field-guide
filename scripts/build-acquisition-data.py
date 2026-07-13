"""Build normalized non-grass acquisition data from the Radical Red v4.1 workbook."""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = next(Path.home().joinpath("Downloads").glob("*Locations & Raid Dens v4.1 - Radical Red.xlsx"))


def text(value):
    return " ".join(str(value or "").replace("_x0010_", "").split())


def level(value):
    if isinstance(value, datetime):
        return f"{value.month}-{value.day}"
    return text(value)


def rarity(value):
    if isinstance(value, (int, float)):
        return round(value * 100, 2)
    return None


def title(value):
    value = text(value).replace("VERMILLION", "VERMILION")
    return value.title().replace("Mt. Moon", "Mt. Moon").replace("S.s.", "S.S.")


def add(target, **entry):
    entry = {key: value for key, value in entry.items() if value not in (None, "")}
    marker = json.dumps(entry, sort_keys=True, ensure_ascii=False)
    if marker not in target["_seen"]:
        target["_seen"].add(marker)
        target["items"].append(entry)


wb = load_workbook(SOURCE, data_only=True, read_only=False)
sections = {name: {"items": [], "_seen": set()} for name in ("wild", "safari", "raids", "special", "gifts", "trades", "fossils", "unobtainable")}

# Fishing and surfing: locations repeat horizontally in five-column blocks.
ws = wb["Fishing & Surfing"]
for method, header_row, rows in (("Old Rod", 3, range(5, 7)), ("Good Rod", 8, range(10, 13)), ("Super Rod", 14, range(16, 21)), ("Surfing", 22, range(24, 29))):
    for col in range(7, ws.max_column + 1, 5):
        place = title(ws.cell(header_row, col).value)
        if not place:
            continue
        for row in rows:
            pokemon = text(ws.cell(row, col + 1).value)
            if pokemon:
                add(sections["wild"], pokemon=pokemon, location=place, method=method, period="any", rarity=rarity(ws.cell(row, col - 1).value), level=level(ws.cell(row, col + 2).value))

# Safari Zone: five areas, each with day, night, rods and surfing columns.
ws = wb["Safari Zone"]
method_columns = (("Grass", "day", 3, 5, 6), ("Grass", "night", 8, 10, 11), ("Old Rod", "any", 13, 15, 16), ("Good Rod", "any", 18, 20, 21), ("Super Rod", "any", 23, 25, 26), ("Surfing", "any", 28, 30, 31))
for start in (2, 19, 36, 53, 70):
    area = title(ws.cell(start, 3).value)
    place = f"Safari Zone — {area}" if area else "Safari Zone"
    for method, period, rarity_col, pokemon_col, level_col in method_columns:
        for row in range(start + 3, min(start + 15, ws.max_row + 1)):
            pokemon = text(ws.cell(row, pokemon_col).value)
            if pokemon and pokemon.lower() != "pokémon":
                add(sections["safari"], pokemon=pokemon, location=place, method=method, period=period, rarity=rarity(ws.cell(row, rarity_col).value), level=level(ws.cell(row, level_col).value))

# Raid Den headers are followed by five possible Pokémon across the next row.
ws = wb["Raid Dens"]
for row in range(1, ws.max_row):
    heading = text(ws.cell(row, 3).value)
    match = re.search(r"--\s*(.*?)\s*--\s*(★+)", heading)
    if not match:
        continue
    place, stars = title(match.group(1)), len(match.group(2))
    for col in (3, 8, 13, 18, 23):
        pokemon = text(ws.cell(row + 1, col).value)
        if pokemon and pokemon.lower() != "pokémon":
            add(sections["raids"], pokemon=pokemon, location=place, method=f"{stars}-Star Raid Den", stars=stars, details="Raid contents may change hourly; use a Wishing Piece to reactivate a cleared den.")

# Static and special encounters. Blank instructions directly following a form inherit the previous instruction.
ws = wb["Statics & Special Pokemon"]
last_method = ""
for row in range(4, ws.max_row + 1):
    pokemon, method = text(ws.cell(row, 4).value), text(ws.cell(row, 6).value)
    if not pokemon or pokemon.upper() == "POKEMON":
        continue
    if method:
        last_method = method
    add(sections["special"], pokemon=pokemon, location="Static & Special", method=method or last_method)

# Gift sheet: location headings are in column C and Pokémon in column D.
ws = wb["Gifts"]
place = "Gift Pokémon"
for row in range(1, ws.max_row + 1):
    heading, pokemon = text(ws.cell(row, 3).value), text(ws.cell(row, 4).value)
    if heading and heading.upper() not in {"POKEMON", "POST-GAME"}:
        place = title(heading)
    if pokemon and pokemon.upper() != "POKEMON":
        details = " · ".join(filter(None, (text(ws.cell(row, 5).value), text(ws.cell(row, 9).value))))
        add(sections["gifts"], pokemon=pokemon, location=place, method="Gift Pokémon", details=details)

# Game Corner purchase list.
for pokemon in ("Dratini", "Larvitar", "Beldum", "Bagon", "Riolu", "Rotom", "Gible", "Larvesta", "Deino", "Goomy", "Jangmo-o", "Toxel", "Dreepy", "Honedge", "Frigibax"):
    add(sections["gifts"], pokemon=pokemon, location="Celadon Game Corner", method="Purchase", cost="₽100,000", details="Hidden Ability is free; a shiny costs an additional ₽100,000.")

# Compact sheets whose layout is primarily prose are kept as explicit source-backed entries.
trade_rows = (
    ("Carnivine", "Cerulean City", "Trade Snom"), ("Chatot", "Celadon Restaurant", "Trade Murkrow"),
    ("Eiscue", "Route 5 Underground", "Trade Carbink"), ("Morpeko", "Route 18 Gate", "Trade Dedenne"),
    ("Farfetch'd-Galar", "Vermilion City", "Trade Pikipek"), ("Mimikyu", "Cinnabar Lab", "Trade Aegislash"),
    ("Floette-Eternal", "Route 11 Gate", "Trade Florges"), ("Ursaluna-BM", "Cinnabar Lab", "Trade Ursaluna"),
)
for pokemon, place, method in trade_rows:
    add(sections["trades"], pokemon=pokemon, location=place, method=method)

fossil_rows = (
    ("Lileep", "Green Shard"), ("Archen", "Green Shard"), ("Omanyte", "Blue Shard or Helix Fossil"),
    ("Kabuto", "Blue Shard or Dome Fossil"), ("Tirtouga", "Blue Shard"), ("Cranidos", "Red Shard"),
    ("Tyrunt", "Red Shard"), ("Anorith", "Yellow Shard"), ("Shieldon", "Yellow Shard"),
    ("Amaura", "Yellow Shard"), ("Aerodactyl", "Old Amber"),
)
for pokemon, method in fossil_rows:
    add(sections["fossils"], pokemon=pokemon, location="Vermilion Fan Club" if "Shard" in method else ("Mt. Moon" if "Fossil" in method else "Pewter Museum"), method=f"Revive/trade: {method}")

ws = wb["Unobtainables"]
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        pokemon = text(ws.cell(row, col).value)
        if pokemon and pokemon not in {"POKEMON", "Pokémon", "Unavailable Pokémon"} and not pokemon.startswith("-"):
            if any(token in pokemon for token in ("Arceus", "Dialga-Origin", "Eternatus", "Magearna-Original", "Palkia-Origin")):
                add(sections["unobtainable"], pokemon=pokemon, location="Unobtainable in v4.1", method="Not obtainable during normal play")

result = {name: value["items"] for name, value in sections.items()}
json_path = ROOT / "data" / "acquisition-data.json"
js_path = ROOT / "data" / "acquisition-data.js"
json_text = json.dumps(result, ensure_ascii=False, indent=2)
json_path.write_text(json_text + "\n", encoding="utf-8")
js_path.write_text("window.ACQUISITION_GUIDE_DATA = " + json_text + ";\n", encoding="utf-8")
print({name: len(items) for name, items in result.items()})
